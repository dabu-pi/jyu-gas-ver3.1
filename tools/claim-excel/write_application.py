#!/usr/bin/env python3
"""
療養費支給申請書 転記スクリプト（openpyxl版）

テンプレートxlsxの書式・結合セルを一切壊さずに、
転記データ（JSON）の値をセルに書き込む。

使い方:
  【一括モード（推奨）】
  1. GASメニュー「一括JSON出力（月指定）」→ DriveにNDJSON出力
  2. DriveからNDJSONファイルをダウンロード
  3. python write_application.py --batch
     → Downloadsフォルダから最新のNDJSONを自動検出して一括転記

  ファイル指定も可能:
  python write_application.py --batch transfer_batch_2026-02.ndjson

  【単一患者モード】
  python write_application.py transfer_data.json
  python write_application.py  (対話モード)
"""

import json
import sys
import os
import glob as glob_mod
import copy
from datetime import datetime, date, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    sys.exit(1)

# ===== 設定 =====
TEMPLATE_FILE = "application_template.xlsx"
TEMPLATE_SHEET = "新　様式第5号"

# ===== 保険者番号: 8桁を1桁ずつ（各4列結合セル） =====
INSURER_NO_CELLS = ["CQ4", "CU4", "CY4", "DC4", "DG4", "DK4", "DO4", "DS4"]

# ===== 保険種別: 該当番号に○付け（参照用: 実際の○書込は SELECTION_SPLIT_MAP 経由）=====
INSURANCE_TYPE_CELLS = {
    1: "CB8",   # 協会けんぽ
    2: "CF8",   # 組合
    3: "CJ8",   # 共済
    4: "CB11",  # 国保
    5: "CF11",  # 退職
    6: "CJ11",  # 後期高齢
}

# ===== 選択肢セル分割マップ（○専用セル方式）=====
# テンプレートの結合セルを出力ファイル内で「ラベル行＋マーカー行」に分割する。
# テンプレートファイルは変更せず、毎回ロードした出力ファイルのみに分割を適用する。
# 書式: {key: (full_merge, label_merge, marker_merge, label_text)}
#   full_merge  : テンプレート元の結合範囲（分割前）
#   label_merge : 上段ラベル行（テキスト保持）
#   marker_merge: 下段マーカー行（選択時「○」を書込む）
#   label_text  : テンプレートのラベル文字列（書き戻し用）
SELECTION_SPLIT_MAP = {
    # --- 性別 (AL21:AO22/AL23:AO24 各4x2 → 4x1 ラベル + 4x1 マーカー) ---
    "gender_男": ("AL21:AO22", "AL21:AO21", "AL22:AO22", "1 男"),
    "gender_女": ("AL23:AO24", "AL23:AO23", "AL24:AO24", "2 女"),
    # --- 保険種別 (各4x3 → 4x2 ラベル + 4x1 マーカー) ---
    "ins_1": ("CB8:CE10",  "CB8:CE9",   "CB10:CE10", "1.協"),
    "ins_2": ("CF8:CI10",  "CF8:CI9",   "CF10:CI10", "2.組"),
    "ins_3": ("CJ8:CM10",  "CJ8:CM9",   "CJ10:CM10", "3.共"),
    "ins_4": ("CB11:CE13", "CB11:CE12", "CB13:CE13", "4.国"),
    "ins_5": ("CF11:CI13", "CF11:CI12", "CF13:CI13", "5.退"),
    "ins_6": ("CJ11:CM13", "CJ11:CM12", "CJ13:CM13", "6.後期"),
    # --- 単独区分 (各6x2 → 6x1 ラベル + 6x1 マーカー) ---
    "tankei_1": ("CT8:CY9",   "CT8:CY8",   "CT9:CY9",   "1.単独"),
    "tankei_2": ("CT10:CY11", "CT10:CY10", "CT11:CY11", "2.2併"),
    "tankei_3": ("CT12:CY13", "CT12:CY12", "CT13:CY13", "3.3併"),
    # --- 本家区分 (各6x2 → 6x1 ラベル + 6x1 マーカー) ---
    "honke_DB8":  ("DB8:DG9",   "DB8:DG8",   "DB9:DG9",   "2.本人"),
    "honke_DB10": ("DB10:DG11", "DB10:DG10", "DB11:DG11", "4.六歳"),
    "honke_DB12": ("DB12:DG13", "DB12:DG12", "DB13:DG13", "6.家族"),
    "honke_DH8":  ("DH8:DM9",   "DH8:DM8",   "DH9:DM9",   "8.高一"),
    "honke_DH12": ("DH12:DM13", "DH12:DM12", "DH13:DM13", "0.高7"),
}

# D4 負傷の原因: BR20:DV24（5行結合）をラベル行(BR20)＋内容行(BR21-24)に分離
# (full_merge, label_merge, content_merge, label_text)
D4_INJURY_CELL_SPLIT = ("BR20:DV24", "BR20:DV20", "BR21:DV24", "負傷の原因")
D4_INJURY_CONTENT_CELL = "BR21"  # 内容書込先（分離後のコンテンツ行左上）

# ===== 英語日付文字列 → 和暦表記 変換（Python側安全網） =====
# GAS 側 Utilities.formatDate が効かなかった場合（String型セルなど）への対策。
# "Mon Feb 02 2026 00:00:00 GMT+0900" → "2026/02/02" → さらに put_wareki_ymd 書式へ変換せず
# _build_injury_text 内でセグメントとして使うため ISO 形式（YYYY/MM/DD）に正規化するだけでよい。
import re as _re

_EN_MONTH = {
    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
    "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
    "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
}

def _normalize_date_str(s: str) -> str:
    """
    JavaScript Date.toString() 形式の英語日付を YYYY/MM/DD に変換する。
    例: "Mon Feb 02 2026 00:00:00 GMT+0900 (JST)" → "2026/02/02"
    該当しない文字列はそのまま返す。
    """
    m = _re.match(
        r"\w{3}\s+(\w{3})\s+(\d{2})\s+(\d{4})", s.strip()
    )
    if m:
        mon_str, day, year = m.group(1), m.group(2), m.group(3)
        mon = _EN_MONTH.get(mon_str)
        if mon:
            return f"{year}/{mon}/{day}"
    return s

# ===== セルマッピング =====
CELL_MAP = {
    "記号番号": "CK5",  # BZ5はテンプレートラベル「記号・番号」を保持

    "被保険者氏名": "X14",
    "住所": "BF14",

    "患者氏名": "E21",
    "生年月日_元号": "AP21",
    "生年月日_年": "AY23",

    "請求区分": "DH31",
    "経過":     "M31",  # D2 継続月数・頻回 補助表示（★正本は摘要欄・頻回欄0.5）
    "摘要": "E44",

    # ===== 施術機関固定情報（全患者共通: NDJSON meta から取得）=====
    "都道府県番号": "CI2",    # U1: CI2:CL3 マージ。施術機関所在都道府県番号（2桁）
    "施術機関コード": "CZ2",  # U2: CZ2:DV3 マージ。登録記号番号の数字部分（★暫定運用）
    # 注: "登録記号番号" は CR49:DV50（ラベル行）のため CELL_MAP から除外。
    #     分割書込先は TOROKU_KIGO_SPLIT_CELLS で管理。
}

# ===== U6 給付割合: 一部負担金割合 → 対象数字上に楕円画像（○専用画像方式）=====
# テンプレート実値: DP8:DV10 = '10・9'（7列×3行）/ DP11:DV13 = '8・7'（7列×3行）
# テンプレート文字列は変更せず、対象数字の上に透明背景楕円画像を浮かべる。
#
# 数字の列方向位置（テンプレートスキャン確認: DP=col120, DV=col126 = 7列）:
#   '10・9'（4文字, 中央揃え）: 左余白1.5列 → '1'@DQ, '0'@DR, '・'@DS, '9'@DT-DU
#   '8・7'  (3文字, 中央揃え)  : 左余白2列   → '8'@DR-DS, '・'@DT, '7'@DT-DU
# ★ズレがある場合はセル範囲を1列ずつ左右にずらして調整すること
KYUFU_OVAL_MAP = {
    1: "DS8:DV10",   # 9割給付: '9' の位置（4列幅: DS-DV, rows 8-10）
    2: "DQ11:DT13",  # 8割給付: '8' の位置（4列幅: DQ-DT, rows 11-13）
    3: "DS11:DV13",  # 7割給付: '7' の位置（4列幅: DS-DV, rows 11-13）
}

# ===== 選択肢楕円の配置範囲マップ =====
# _write_selection_marker が参照する配置範囲（TwoCellAnchor 画像）。
# 行番号を1つ下げる（＝上へ移動）ことで、楕円の中央位置を1セル分上にずらせる。
# 給付割合は KYUFU_OVAL_MAP で別管理（横幅優先のため）。
#
# 微調整履歴:
#   v1 (Revision 00012-7mn): marker_merge → full_merge に変更（上N行分移動）
#   v2 (Revision 00013-xtd): full_merge をそのまま使用
#   v3 (Revision 00014-*): 全エントリ -1行（目視で全欄「さらに上1セル」要望）
#   v4 (Revision 00016-*): _apply_selection_splits の選択肢分割を廃止。
#     テンプレ結合を維持するため SELECTION_OVAL_MAP をテンプレ full_merge 値に戻す。
#     【根本原因】分割でテキスト位置が上シフト → 楕円とのズレが発生していた。
SELECTION_OVAL_MAP = {
    # 性別（テンプレ full_merge そのまま）
    "gender_男": "AL21:AO22",
    "gender_女": "AL23:AO24",
    # 保険種別（テンプレ full_merge そのまま）
    "ins_1": "CB8:CE10",
    "ins_2": "CF8:CI10",
    "ins_3": "CJ8:CM10",
    "ins_4": "CB11:CE13",
    "ins_5": "CF11:CI13",
    "ins_6": "CJ11:CM13",
    # 単独区分（テンプレ full_merge そのまま）
    "tankei_1": "CT8:CY9",
    "tankei_2": "CT10:CY11",
    "tankei_3": "CT12:CY13",
    # 本家区分（テンプレ full_merge そのまま）
    "honke_DB8":  "DB8:DG9",
    "honke_DB10": "DB10:DG11",
    "honke_DB12": "DB12:DG13",
    "honke_DH8":  "DH8:DM9",
    "honke_DH12": "DH12:DM13",
}

# ===== 楕円スタイル =====
# "normal": 通常の選択肢欄（性別/保険種別/単独/本家）
# "kyufu":  給付割合専用（縦方向マージンを小さくして横長に見せる）
OVAL_STYLES = {
    "normal": {"margin_emu": 19050, "line_width": 2},
    "kyufu":  {"margin_emu": 9525,  "line_width": 2},
}

# ===== 下段 登録記号番号 分割欄（行51-52）=====
# CR49:DV50 はラベル行「登録記号番号」→ 書き込み禁止
# 入力欄: 左=CR51:DH52 / 中=DK51:DO52 / 右=DR51:DV52
# 区切りセル DI51:DJ52 / DP51:DQ52 は触らない（テンプレートのハイフン表示用）
TOROKU_KIGO_SPLIT_CELLS = ("CR51", "DK51", "DR51")

# ===== 初検料・再検料・計: ラベル内に金額を埋め込む =====
# セル構造: E33:X33 = "初検料　　　　　　　　円" のようにラベル内テキスト
LABEL_AMOUNT_CELLS = {
    "初検料":         {"cell": "E33",  "tmpl": "初検料{amt}円"},
    "初検時相談支援料": {"cell": "Y33",  "tmpl": "初検時相談\n支援料{amt}円"},
    "再検料":         {"cell": "Y34",  "tmpl": "再検料{amt}円"},
    "基本3項目_計":   {"cell": "DC33", "tmpl": "{amt}円"},
}

# ===== 施療料: ラベル内に金額埋め込み =====
# AC35:AQ35 = "(1)　　　　　　　円" → "(1) XXXX 円"
SHORYO_CELLS = [
    {"cell": "AC35", "no": 1},
    {"cell": "AR35", "no": 2},
    {"cell": "BG35", "no": 3},
    {"cell": "BV35", "no": 4},
    {"cell": "CK35", "no": 5},
]
SHORYO_TOTAL_CELL = {"cell": "DC35", "tmpl": "{amt}円"}

# ===== 合計欄: 1桁ずつ（各4列結合セル×6 + 末尾"円"） =====
AMOUNT_DIGIT_CELLS = {
    "合計":     ["CV44", "CZ44", "DD44", "DH44", "DL44", "DP44"],
    "一部負担金": ["CV45", "CZ45", "DD45", "DH45", "DL45", "DP45"],
    "請求金額":   ["CV46", "CZ46", "DD46", "DH46", "DL46", "DP46"],
}

# ===== 負傷名行（最大5行: 行26-30）=====
# 年月日はすべて和暦年・月・日の3セル分割
INJURY_ROWS = [
    {"row": 26, "name": "E26", "injY": "AN26", "injM": "AS26", "injD": "AY26",
     "iniY": "BD26", "iniM": "BI26", "iniD": "BO26",
     "stY": "BT26", "stM": "BY26", "stD": "CE26",
     "edY": "CJ26", "edM": "CO26", "edD": "CU26",
     "days": "CZ26", "tenki": "DH26"},
    {"row": 27, "name": "E27", "injY": "AN27", "injM": "AS27", "injD": "AY27",
     "iniY": "BD27", "iniM": "BI27", "iniD": "BO27",
     "stY": "BT27", "stM": "BY27", "stD": "CE27",
     "edY": "CJ27", "edM": "CO27", "edD": "CU27",
     "days": "CZ27", "tenki": "DH27"},
    {"row": 28, "name": "E28", "injY": "AN28", "injM": "AS28", "injD": "AY28",
     "iniY": "BD28", "iniM": "BI28", "iniD": "BO28",
     "stY": "BT28", "stM": "BY28", "stD": "CE28",
     "edY": "CJ28", "edM": "CO28", "edD": "CU28",
     "days": "CZ28", "tenki": "DH28"},
    {"row": 29, "name": "E29", "injY": "AN29", "injM": "AS29", "injD": "AY29",
     "iniY": "BD29", "iniM": "BI29", "iniD": "BO29",
     "stY": "BT29", "stM": "BY29", "stD": "CE29",
     "edY": "CJ29", "edM": "CO29", "edD": "CU29",
     "days": "CZ29", "tenki": "DH29"},
    {"row": 30, "name": "E30", "injY": "AN30", "injM": "AS30", "injD": "AY30",
     "iniY": "BD30", "iniM": "BI30", "iniD": "BO30",
     "stY": "BT30", "stM": "BY30", "stD": "CE30",
     "edY": "CJ30", "edM": "CO30", "edD": "CU30",
     "days": "CZ30", "tenki": "DH30"},
]

# ===== 部位別明細行 =====
PART_ROWS = [
    {"row": 38, "label": "E38", "teiRate": "H38", "teiStart": "M38",
     "koryoUnit": "V38", "koryoCnt": "AC38", "koryoAmt": "AH38",
     "coldCnt": "AR38", "coldAmt": "AW38",
     "warmCnt": "BF38", "warmAmt": "BK38",
     "elecCnt": "BT38", "elecAmt": "BY38",
     "subtotal": "CH38", "multiCoef": "CR38", "multiTotal": "CW38",
     "longCoef": "DF38", "longTotal": "DK38"},
    {"row": 39, "label": "E39", "teiRate": "H39", "teiStart": "M39",
     "koryoUnit": "V39", "koryoCnt": "AC39", "koryoAmt": "AH39",
     "coldCnt": "AR39", "coldAmt": "AW39",
     "warmCnt": "BF39", "warmAmt": "BK39",
     "elecCnt": "BT39", "elecAmt": "BY39",
     "subtotal": "CH39", "multiCoef": "CR39", "multiTotal": "CW39",
     "longCoef": "DF39", "longTotal": "DK39"},
    {"row": 40, "label": "E40", "teiRate": "H40", "teiStart": "M40",
     "koryoUnit": "V40", "koryoCnt": "AC40", "koryoAmt": "AH40",
     "coldCnt": "AR40", "coldAmt": "AW40",
     "warmCnt": "BF40", "warmAmt": "BK40",
     "elecCnt": "BT40", "elecAmt": "BY40",
     "subtotal": "CH40", "multiCoef": "CR40", "multiTotal": "CW40",
     "longCoef": "DF40", "longTotal": "DK40"},
    {"row": 42, "label": "E42", "teiRate": "H42", "teiStart": "M42",
     "koryoUnit": "V42", "koryoCnt": "AC42", "koryoAmt": "AH42",
     "coldCnt": "AR42", "coldAmt": "AW42",
     "warmCnt": "BF42", "warmAmt": "BK42",
     "elecCnt": "BT42", "elecAmt": "BY42",
     "subtotal": "CH42", "multiCoef": "CR42", "multiTotal": "CW42",
     "longCoef": "DF42", "longTotal": "DK42"},
]


# ===== 和暦変換 =====
def to_wareki(d: date):
    """西暦→和暦 (code, year)  code: 1=明治 2=大正 3=昭和 4=平成 5=令和"""
    if d >= date(2019, 5, 1):
        return 5, d.year - 2018
    if d >= date(1989, 1, 8):
        return 4, d.year - 1988
    if d >= date(1926, 12, 25):
        return 3, d.year - 1925
    if d >= date(1912, 7, 30):
        return 2, d.year - 1911
    return 1, d.year - 1867


def parse_date(val):
    """文字列/日付 → date or None"""
    if val is None or val == "":
        return None
    if isinstance(val, (datetime, date)):
        return val if isinstance(val, date) else val.date()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def safe_num(val):
    """数値変換（空/非数値→None）"""
    if val is None or val == "":
        return None
    try:
        n = float(val)
        if n == 0:
            return None
        return n
    except (ValueError, TypeError):
        return None


def safe_int(val):
    """整数変換（空/非数値→None, 0→None）"""
    n = safe_num(val)
    if n is None:
        return None
    return int(n)


# ===== 転帰：楕円画像で丸囲み =====
def _add_tenki_oval(ws, tenki_cell: str, tenki_value: str):
    """
    転帰セル（例: DH26）の「治癒・中止・転医」テキスト上に、
    該当する語句を透明背景の楕円PNG画像で囲む。

    openpyxl は独自Shape描画の永続化をサポートしないため、
    PIL で透明楕円PNGを生成し ws.add_image() で貼り付ける。
    tenki_value: "治癒" / "中止" / "転医"
    """
    from PIL import Image as PILImage, ImageDraw
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    from openpyxl.utils import column_index_from_string
    import re

    # テキスト「治癒・中止・転医」の中で、各語句の列位置を直接指定
    # 結合セル DH:DV (15列)、テキスト中央揃え
    # テンプレート列幅: 全列 width=1 (1文字単位)
    # 「治癒・中止・転医」=全角8文字、MSP明朝10.5pt中央揃え
    # 左余白 = (15-8)/2 = 3.5列、各文字 ≈1列幅
    #   治癒: DH+3.5～5.5、中止: DH+6.5～8.5、転医: DH+9.5～11.5
    # 楕円の開始/終了列（DH起点のオフセット、列単位、0.5列マージン付き）
    tenki_col_offsets = {
        "治癒": (3, 6),     # DH+3 ～ DH+6  (治: 3.5-4.5, 癒: 4.5-5.5)
        "中止": (6, 9),     # DH+6 ～ DH+9  (中: 6.5-7.5, 止: 7.5-8.5)
        "転医": (9, 12),    # DH+9 ～ DH+12 (転: 9.5-10.5, 医: 10.5-11.5)
    }
    if tenki_value not in tenki_col_offsets:
        return

    off_start, off_end = tenki_col_offsets[tenki_value]

    # セル座標をパース（例: "DH26" → col=DH, row=26）
    m_cell = re.match(r"([A-Z]+)(\d+)", tenki_cell)
    if not m_cell:
        return
    start_col_letter = m_cell.group(1)
    row_0 = int(m_cell.group(2)) - 1  # 0-based row

    dh_col_0 = column_index_from_string(start_col_letter) - 1  # 0-based

    # 1列 width=1 ≈ 12px → EMU = 12 * 9525 = 114300
    col_width_emu = 114300

    oval_start = dh_col_0 + off_start
    oval_end = dh_col_0 + off_end

    from_col = int(oval_start)
    from_col_off = int((oval_start - from_col) * col_width_emu)
    to_col = int(oval_end)
    to_col_off = int((oval_end - to_col) * col_width_emu)

    # 行方向: セルの上下に少し余白（行高 30.75pt = 390525 EMU）
    row_height_emu = 390525
    v_margin = int(row_height_emu * 0.10)

    # --- 透明背景に黒楕円の輪郭だけ描いた PNG を生成 ---
    img_w, img_h = 240, 80
    pil_img = PILImage.new('RGBA', (img_w, img_h), (0, 0, 0, 0))
    draw = ImageDraw.Draw(pil_img)
    draw.ellipse([2, 2, img_w - 3, img_h - 3], outline='black', width=2)

    # BytesIO 経由で openpyxl Image を作成（直接PIL渡しだと fp 属性がない）
    from io import BytesIO
    buf = BytesIO()
    pil_img.save(buf, format='PNG')
    buf.seek(0)
    xl_img = XlImage(buf)
    anchor = TwoCellAnchor(
        _from=AnchorMarker(
            col=from_col, colOff=from_col_off,
            row=row_0, rowOff=v_margin,
        ),
        to=AnchorMarker(
            col=to_col, colOff=to_col_off,
            row=row_0, rowOff=row_height_emu - v_margin,
        ),
        editAs='oneCell',
    )
    xl_img.anchor = anchor
    ws.add_image(xl_img)


# ===== ○専用画像方式: 汎用楕円アンカー =====

def _draw_oval_on_range(ws, cell_range: str, style: str = "normal"):
    """
    指定セル範囲 "A1:C3" に透明背景の楕円 PNG 画像を TwoCellAnchor でアンカー配置する。

    テキスト "○" の代わりに使うことで、列幅・結合セル・フォントサイズに依存しない
    視覚的中央配置を実現する。配置基準は SELECTION_OVAL_MAP / KYUFU_OVAL_MAP の
    セル範囲文字列で管理し、微調整時はマップの値を1列ずらすだけでよい。

    転帰 (_add_tenki_oval) と同じ PIL + TwoCellAnchor 方式を汎用化したもの。

    AnchorMarker の col/row は 0-based:
      from_col = column_index_from_string(start_col) - 1  ← left edge of start col
      to_col   = column_index_from_string(end_col)        ← right edge of end col
      from_row = start_row - 1                            ← top of start row
      to_row   = end_row                                  ← bottom of end row

    style: OVAL_STYLES のキー。"normal"（通常選択欄）/ "kyufu"（給付割合専用）。
    """
    from PIL import Image as PILImage, ImageDraw
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    from openpyxl.utils import column_index_from_string
    from io import BytesIO
    import re

    oval_style = OVAL_STYLES.get(style, OVAL_STYLES["normal"])
    margin_emu = oval_style["margin_emu"]
    line_width = oval_style["line_width"]

    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", cell_range.strip())
    if not m:
        return
    sc, sr_str, ec, er_str = m.group(1), m.group(2), m.group(3), m.group(4)
    sr, er = int(sr_str), int(er_str)

    # AnchorMarker 座標 (0-based)
    from_col = column_index_from_string(sc) - 1   # left edge of sc
    from_row = sr - 1                              # top edge of sr
    to_col   = column_index_from_string(ec)        # right edge of ec (= left of ec+1)
    to_row   = er                                  # bottom edge of er (= top of er+1)

    # PNG サイズ: TwoCellAnchor で引き伸ばされるためアスペクト比を大まかに合わせる
    col_span = column_index_from_string(ec) - column_index_from_string(sc) + 1
    row_span = er - sr + 1
    img_h = max(20, row_span * 20)
    img_w = max(30, col_span * 15)

    # 透明背景 + 黒楕円輪郭 PNG を生成
    pil_img = PILImage.new('RGBA', (img_w, img_h), (0, 0, 0, 0))
    draw = ImageDraw.Draw(pil_img)
    draw.ellipse([1, 1, img_w - 2, img_h - 2], outline='black', width=line_width)

    buf = BytesIO()
    pil_img.save(buf, format='PNG')
    buf.seek(0)

    xl_img = XlImage(buf)
    xl_img.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=from_col, colOff=margin_emu, row=from_row, rowOff=margin_emu),
        to=AnchorMarker(col=to_col,   colOff=0,           row=to_row,   rowOff=0),
        editAs='oneCell',
    )
    ws.add_image(xl_img)


# ===== D4 負傷の原因セル分割 =====

def _apply_selection_splits(ws):
    """
    D4 負傷の原因（BR20:DV24）を「ラベル行 BR20 + 内容行 BR21-24」に分割し、
    ラベル「負傷の原因」を BR20:DV20 に書き戻す。

    ★ 選択肢セル（性別/保険種別/単独区分/本家区分）の分割は行わない。
      - 旧方式: unmerge_cells + re-merge でラベル行＋マーカー行に分割していた
      - 廃止理由: 分割するとテキスト表示位置が上シフトし、楕円画像とのズレが発生する
        （テンプレ: 3行結合の中央 ≈ 行9 → 分割後: 2行ラベル中央 ≈ 行8-9 上よりに移動）
      - 現在の画像オーバーレイ方式では分割不要。テンプレ結合をそのまま維持する。

    テンプレートファイルは変更しない（毎回ロードされる）。
    """
    # D4: BR20:DV24 → ラベル行 BR20:DV20 + 内容行 BR21:DV24
    full, label, content, text = D4_INJURY_CELL_SPLIT
    ws.unmerge_cells(full)
    ws.merge_cells(label)
    ws[label.split(':')[0]] = text
    ws.merge_cells(content)


def _write_selection_marker(ws, marker_key: str):
    """
    指定キーの選択肢セル全体（ラベル行＋マーカー行）に楕円画像を配置する（○専用画像方式）。
    _apply_selection_splits() が事前に呼ばれていることが前提。

    テキスト "○" は書き込まない。PIL TwoCellAnchor で画像がセル中央に浮かぶ。
    配置範囲は SELECTION_OVAL_MAP（full_merge）を使用するため、マーカー行のみを
    指定した旧方式に対して「上N行分」移動した位置に楕円が描かれる。
      ・保険種別: 3行全体 → 実質「上2行」移動
      ・性別/単独/本家: 2行全体 → 実質「上1行」移動
    位置微調整は SELECTION_OVAL_MAP の範囲文字列を変更して行う。
    """
    if marker_key not in SELECTION_OVAL_MAP:
        return
    oval_range = SELECTION_OVAL_MAP[marker_key]
    _draw_oval_on_range(ws, oval_range, style="normal")


# ===== メイン転記 =====
def write_application(template_path: str, json_data: dict, output_path: str, clinic_info: dict = None):
    """
    テンプレートxlsxを開き、転記データを書き込んで別名保存。
    書式・結合は保持される。

    json_data: { "case1": {...}, "case2": {...} | null }
    clinic_info: { "prefectureNo": "14", "torokuKigoNo": "契2804440-0-0" }
      全患者共通の施術機関固定情報。None の場合は書込をスキップ（後方互換）。
      NDJSON の _meta 行から取得し、batch_write / batch_write_from_string が渡す。
    """
    wb = load_workbook(template_path)
    ws = wb[TEMPLATE_SHEET]

    # 選択肢セルを「ラベル行＋マーカー行」に分割（○専用セル方式の前処理）
    _apply_selection_splits(ws)

    count = 0
    row1 = json_data.get("case1") or {}
    row2 = json_data.get("case2")

    def put(cell, val):
        nonlocal count
        if val is None or val == "":
            return
        ws[cell] = val
        count += 1

    def put_num(cell, val):
        nonlocal count
        n = safe_num(val)
        if n is None:
            return
        ws[cell] = n
        count += 1

    # ===== タイトル行: 「令和　　年　　月分」に対象月を埋め込む =====
    target_month = str(row1.get("対象月", "") or "").strip()
    if target_month:
        ym_parts = target_month.split("-")
        if len(ym_parts) >= 2:
            ym_date = date(int(ym_parts[0]), int(ym_parts[1]), 1)
            _code, w_year = to_wareki(ym_date)
            month_num = int(ym_parts[1])
            title = ws["E3"].value
            if title:
                ws["E3"] = title.replace("令和　　年　　月分", f"令和　{w_year}年　{month_num}月分")
                count += 1

    # --- 修正①: 保険者番号を1桁ずつ ---
    insurer_no = str(row1.get("保険者番号", "") or "").strip()
    if insurer_no:
        # 8桁に左ゼロ詰め
        insurer_no = insurer_no.zfill(8)
        for i, cell in enumerate(INSURER_NO_CELLS):
            if i < len(insurer_no):
                ws[cell] = int(insurer_no[i])
                count += 1

    # 記号・番号を結合してCK5に入力（BZ5のラベルは保持）
    symbol = str(row1.get("記号", "") or "").strip()
    number = str(row1.get("番号", "") or "").strip()
    if symbol or number:
        put(CELL_MAP["記号番号"], f"{symbol}・{number}")

    # ===== 保険種別○付け（○専用セル方式）=====
    # マスタ値優先、なければ保険者番号から自動判定
    # テンプレセル（1.協 等）は変更せず、下段マーカー行に「○」を書込む
    ins_type = safe_int(row1.get("保険種別"))
    if ins_type is None:
        ins_type = detect_insurance_type(insurer_no)
    if ins_type and 1 <= ins_type <= 6:
        _write_selection_marker(ws, f"ins_{ins_type}")
        count += 1

    # ===== 被保険者 =====
    put(CELL_MAP["被保険者氏名"], row1.get("被保険者氏名"))
    put(CELL_MAP["住所"], row1.get("住所"))

    # ===== 受療者 =====
    put(CELL_MAP["患者氏名"], row1.get("患者氏名"))

    # ===== 性別（○専用セル方式）=====
    # テンプレセル（1 男 / 2 女）は変更せず、下段マーカー行に「○」を書込む
    # 根拠: テンプレートスキャン確認（2026-03-20）
    gender_val = str(row1.get("性別") or "").strip()
    if gender_val in ("男", "女"):
        _write_selection_marker(ws, f"gender_{gender_val}")
        count += 1

    # 生年月日 → 元号○付け + 年月日テキスト
    bd = parse_date(row1.get("患者生年月日"))
    if bd:
        code, year = to_wareki(bd)
        # AP21: 元号テキスト内の該当番号に○（丸数字に置換）
        put_era_circle(ws, "AP21", code)
        count += 1
        # AY23: "   年 　 月　　 日" → "XX年 MM月  DD日"
        ws["AY23"] = f"{year:>3}年 {bd.month:>2}月  {bd.day:>2}日"
        count += 1

    # ===== 負傷名(1)-(5) =====
    # --- 修正②: 年月日を和暦で書き込む ---
    inj_data = build_injury_rows(row1, row2)
    inj_data = [d for d in inj_data if d.get("name")]  # ★空行を除外して詰める
    for i, d in enumerate(inj_data):
        if i >= len(INJURY_ROWS):
            break
        m = INJURY_ROWS[i]
        # 行26（1行目）→（1）、行27（2行目）→（2）を先頭に付ける
        # 行28以降はテンプレートに (3)(4)(5) が既存のため番号不要
        name = d["name"]
        if i == 0:
            name = f"（1）{name}"
        elif i == 1:
            name = f"（2）{name}"
        put(m["name"], name)
        put_wareki_ymd(ws, m["injY"], m["injM"], m["injD"], d.get("injuryDate"))
        put_wareki_ymd(ws, m["iniY"], m["iniM"], m["iniD"], d.get("firstDate"))
        put_wareki_ymd(ws, m["stY"], m["stM"], m["stD"], d.get("startDate"))
        put_wareki_ymd(ws, m["edY"], m["edM"], m["edD"], d.get("endDate"))
        put_num(m["days"], d.get("days"))
        # 転帰（治癒・中止・転医のうち該当語句を楕円図形で丸囲み）
        tenki = str(d.get("tenki", "") or "").strip()
        if tenki and m.get("tenki"):
            _add_tenki_oval(ws, m["tenki"], tenki)
            count += 1
        # 日付書き込みカウント
        for dk in ("injuryDate", "firstDate", "startDate", "endDate"):
            if d.get(dk):
                count += 3

    # --- 修正③: 初検料/再検料/計 ラベル内金額埋め込み ---
    put_label_amount(ws, LABEL_AMOUNT_CELLS["初検料"], row1.get("初検料_月額"))
    put_label_amount(ws, LABEL_AMOUNT_CELLS["初検時相談支援料"], row1.get("初検時相談支援料_月額"))
    put_label_amount(ws, LABEL_AMOUNT_CELLS["再検料"], row1.get("再検料_月額"))
    put_label_amount(ws, LABEL_AMOUNT_CELLS["基本3項目_計"], row1.get("基本3項目_計"))
    count += sum(1 for k in ("初検料_月額", "初検時相談支援料_月額", "再検料_月額", "基本3項目_計")
                 if safe_num(row1.get(k)) is not None)

    # ===== 施療料: ラベル内金額 =====
    # Fix-S: 非ゼロ値のみに詰めてから連番で書き込む（傷病名と同じ詰め寄せ方式）
    shoryo_data = [v for v in build_shoryo_array(row1, row2) if safe_int(v) is not None]
    for i, val in enumerate(shoryo_data):
        if i >= len(SHORYO_CELLS):
            break
        sc = SHORYO_CELLS[i]
        ws[sc["cell"]] = f"({sc['no']}){safe_int(val):>10,}円"
        count += 1

    # 施療料計
    shoryo_total = int((safe_num(row1.get("施療料_計")) or 0) +
                       (safe_num((row2 or {}).get("施療料_計")) or 0))
    if shoryo_total > 0:
        ws[SHORYO_TOTAL_CELL["cell"]] = f"{shoryo_total:,}円"
        count += 1

    # ===== 部位別明細 =====
    # Fix-P: has_data=True の行だけ display_idx を進め、PART_ROWS と labels を詰めて参照する
    part_data = build_part_detail_array(row1, row2)
    labels = "⑴⑵⑶⑷⑸"
    display_idx = 0
    for d in part_data:
        if display_idx >= len(PART_ROWS):
            break
        if not d.get("has_data"):
            continue
        m = PART_ROWS[display_idx]
        put(m["label"], labels[display_idx] if display_idx < len(labels) else "")
        display_idx += 1
        put_num(m["teiRate"], d.get("tei_rate"))
        put_num(m["koryoUnit"], d.get("koryo_unit"))
        put_num(m["koryoCnt"], d.get("koryo_cnt"))
        put_num(m["koryoAmt"], d.get("koryo_amt"))
        put_num(m["coldCnt"], d.get("cold_cnt"))
        put_num(m["coldAmt"], d.get("cold_amt"))
        put_num(m["warmCnt"], d.get("warm_cnt"))
        put_num(m["warmAmt"], d.get("warm_amt"))
        put_num(m["elecCnt"], d.get("elec_cnt"))
        put_num(m["elecAmt"], d.get("elec_amt"))
        # DK38 = 部位の総額（後療+冷+温+電）
        put_num(m["longTotal"], d.get("subtotal"))

    # ===== 施術日カレンダー M32 =====
    visit_days = json_data.get("visitDays") or []
    if visit_days:
        put_calendar_circles(ws, "M32", visit_days)
        count += 1

    # --- 修正⑤: 合計/一部負担金/請求金額を桁分割 ---
    put_amount_digits(ws, AMOUNT_DIGIT_CELLS["合計"], row1.get("当月合計"))
    put_amount_digits(ws, AMOUNT_DIGIT_CELLS["一部負担金"], row1.get("窓口負担額"))
    put_amount_digits(ws, AMOUNT_DIGIT_CELLS["請求金額"], row1.get("請求金額"))
    for k in ("当月合計", "窓口負担額", "請求金額"):
        n = safe_int(row1.get(k))
        if n is not None:
            count += len(str(abs(n)))

    # ===== U7 請求区分 DH31 =====
    # GAS側 row["請求区分"] = "新規" | "継続" | ""
    # 同月内治癒再発（"新規・継続" 両方○）は将来対応。現時点では文字列をそのまま書く。
    seikyu_kubun = row1.get("請求区分") or ""
    if seikyu_kubun:
        put(CELL_MAP["請求区分"], seikyu_kubun)

    # ===== D2 継続月数・頻回: M31 経過欄（当面未使用・出力停止）=====
    # ★設計確定（2026-03-20）: GAS側が row["経過"]="" を送るためこのブロックは自然にスキップ
    # ★正本=摘要欄（手動）+長期欄（頻回→0.5/長期のみ→0.75、手動）
    # ★将来 M31 自動出力を復活させる場合はコメントアウトを解除すること
    keizoku = str(row1.get("経過") or "").strip()
    if keizoku:  # GAS側が "" を送るため現状は常にスキップ（出力停止中）
        put(CELL_MAP["経過"], keizoku)

    # ===== U5 本家区分 行8-13（○専用セル方式）=====
    # 判定ソース: 保険種別・続柄・生年月日・一部負担金割合・対象月
    # テンプレセル（2.本人 等）は変更せず、下段マーカー行に「○」を書込む
    # 根拠: docs/JREC-01_申請書様式運用メモ.md §4 U5 参照
    honkeku_cell = derive_honkeku_cell(row1)
    if honkeku_cell:
        _write_selection_marker(ws, f"honke_{honkeku_cell}")
        count += 1

    # ===== U6 給付割合 行8-13（○専用画像方式）=====
    # テンプレート文字列（'10・9' / '8・7'）は変更せず、対象数字の上に楕円画像を配置。
    # KYUFU_OVAL_MAP でサブセル範囲を指定。ズレがある場合は1列ずらして調整可能。
    # 根拠: docs/JREC-01_申請書様式運用メモ.md §4 U6 参照
    burden_digit = safe_int(row1.get("一部負担金割合")) or 0
    kyufu_oval_range = KYUFU_OVAL_MAP.get(burden_digit)
    if kyufu_oval_range:
        _draw_oval_on_range(ws, kyufu_oval_range, style="kyufu")
        count += 1

    # ===== D4 負傷の原因（BR21: 分離後のコンテンツ行）=====
    # 出力条件: row2の部位1に金額あり = 申請書3部位目が存在
    #           = 「3部位目を100分の60で算定することとなる場合」
    # 根拠: 柔整療養費告示 別表第2 備考2「3部位目は所定料金の100分の60」
    # セル構造: _apply_selection_splits() により
    #   BR20:DV20 = ラベル行「負傷の原因」（固定）
    #   BR21:DV24 = コンテンツ行（ここに書込む）
    # ★ 暫定ルール: 3部位目の存在を「row2["部位1_計"] > 0」で判定
    # ソース: 「負傷の日時」「負傷の場所」「負傷の状況」（初検情報履歴シート由来）
    part3_has_data = (
        row2 is not None and (
            (safe_num(row2.get("部位1_計")) or 0) > 0 or
            (safe_num(row2.get("部位1_後療料_金額")) or 0) > 0
        )
    )
    if part3_has_data:
        def _build_injury_text(row):
            segs = []
            for k in ("負傷の日時", "負傷の場所", "負傷の状況"):  # 日時→場所→状況
                v = str(row.get(k) or "").strip()
                if v:
                    # 英語日付文字列（GAS String型セル由来）をYYYY/MM/DDに正規化
                    v = _normalize_date_str(v)
                    segs.append(v)
            return "\u3000".join(segs)  # 全角スペース区切り
        t1 = _build_injury_text(row1)
        t2 = _build_injury_text(row2) if row2 else ""
        unique_texts = []
        seen: set = set()
        for t in [t1, t2]:
            if t and t not in seen:
                seen.add(t)
                unique_texts.append(t)
        # 1ケースにつき1行・行頭に（1）（2）番号付き・改行区切り
        injury_lines = [f"（{idx + 1}）{t}" for idx, t in enumerate(unique_texts)]
        injury_text = "\n".join(injury_lines)
        if injury_text:
            # BR21: コンテンツ行（_apply_selection_splits で分離済み）
            ws[D4_INJURY_CONTENT_CELL] = injury_text
            count += 1

    # ===== 施術機関固定情報（clinic_info から取得: 全患者共通）=====
    # U1 都道府県番号 → CI2 / U2 施術機関コード → CZ2 / U4 単独 → CT9（マーカー行）
    # 下段登録記号番号 → CR51(左)/DK51(中)/DR51(右) 分割書込
    # clinic_info が None の場合はスキップ（後方互換: 旧NDJSON/単体テスト対応）
    if clinic_info:
        # U1: 都道府県番号 → CI2
        pref_no = str(clinic_info.get("prefectureNo") or "").strip()
        if pref_no:
            put(CELL_MAP["都道府県番号"], pref_no)

        # U2: 施術機関コード → CZ2（登録記号番号から数字部分を導出: 暫定運用）
        clinic_code = derive_clinic_code(str(clinic_info.get("torokuKigoNo") or ""))
        if clinic_code:
            put(CELL_MAP["施術機関コード"], clinic_code)

        # U4: 単併区分（○専用セル方式）固定「単独」→ tankei_1 マーカー行に○
        _write_selection_marker(ws, "tankei_1")

        # 下段 登録記号番号 → CR51/DK51/DR51（分割書込）
        # CR49:DV50 はラベル行「登録記号番号」→ 書き込まない
        # 例: '契2804440-0-0' → 左='契2804440' / 中='0' / 右='0'
        toroku = str(clinic_info.get("torokuKigoNo") or "").strip()
        if toroku:
            parts = toroku.split("-")
            for i, cell_addr in enumerate(TOROKU_KIGO_SPLIT_CELLS):
                ws[cell_addr] = parts[i] if i < len(parts) else ""

    wb.save(output_path)
    print(f"書込完了: {output_path} ({count}セル)")
    return count


# ===== 書き込みヘルパー =====

def put_wareki_ymd(ws, cell_y, cell_m, cell_d, date_val):
    """
    修正②: 日付を和暦年・月・日で3セルに分解して書き込む。
    年は和暦年（例: 令和8年→8）
    """
    d = parse_date(date_val)
    if d is None:
        return
    _code, wareki_year = to_wareki(d)
    ws[cell_y] = wareki_year
    ws[cell_m] = d.month
    ws[cell_d] = d.day


def put_label_amount(ws, config, val):
    """
    修正③: ラベルセルに金額を埋め込む。
    例: "初検料　　　　　　　　円" → "初検料1,460円"
    """
    n = safe_int(val)
    if n is None:
        return
    text = config["tmpl"].replace("{amt}", f"{n:,}")
    ws[config["cell"]] = text


def put_amount_digits(ws, cells, val):
    """
    修正⑤: 金額を右詰めで1桁ずつ書き込む。
    cells: ["CV44","CZ44","DD44","DH44","DL44","DP44"] の6セル
    例: 18730 → [空,1,8,7,3,0]
    """
    n = safe_int(val)
    if n is None:
        return
    digits = str(abs(n))
    # 6桁に右詰め（6桁未満は左を空ける）
    padded = digits.rjust(len(cells))
    for i, cell in enumerate(cells):
        if i < len(padded) and padded[i] != " ":
            ws[cell] = int(padded[i])


# 丸数字マップ（1〜31）
CIRCLED_NUMBERS = {
    1: "①", 2: "②", 3: "③", 4: "④", 5: "⑤",
    6: "⑥", 7: "⑦", 8: "⑧", 9: "⑨", 10: "⑩",
    11: "⑪", 12: "⑫", 13: "⑬", 14: "⑭", 15: "⑮",
    16: "⑯", 17: "⑰", 18: "⑱", 19: "⑲", 20: "⑳",
    21: "㉑", 22: "㉒", 23: "㉓", 24: "㉔", 25: "㉕",
    26: "㉖", 27: "㉗", 28: "㉘", 29: "㉙", 30: "㉚",
    31: "㉛",
}


def put_era_circle(ws, cell, era_code):
    """
    AP21の元号テキスト内で該当番号を丸数字に置換。
    元テキスト: "1 明 　　2 大　　3 昭　　 4 平　　5令"
    例: era_code=5(令和) → "1 明 　　2 大　　3 昭　　 4 平　　⑤令"
    """
    original = ws[cell].value
    if not original:
        return
    # 元号番号の文字を丸数字に置換
    circled = CIRCLED_NUMBERS.get(era_code, str(era_code))
    target = str(era_code)
    # テキスト内の該当数字を丸数字に置換（最初の1箇所のみ）
    new_text = original.replace(target, circled, 1)
    ws[cell] = new_text


def detect_insurance_type(insurer_no: str) -> int | None:
    """
    保険者番号（8桁）の先頭2桁（法別番号）から保険種別を自動判定。
    返り値: 1〜6 or None（判定不能）

    1: 協会けんぽ（01〜04）
    2: 組合（06, 63〜69）
    3: 共済（31〜34）
    4: 国保（6桁以下＝法別番号なし）
    5: 退職（67）
    6: 後期高齢（39）
    """
    if not insurer_no:
        return None

    # 6桁以下 → 国保（法別番号なし）
    if len(insurer_no) <= 6:
        return 4

    # 先頭2桁 = 法別番号
    houbestu = insurer_no[:2]

    try:
        h = int(houbestu)
    except ValueError:
        return None

    if 1 <= h <= 4:
        return 1   # 協会けんぽ
    if h == 6:
        return 2   # 組合
    if h == 67:
        return 5   # 退職
    if 63 <= h <= 69:
        return 2   # 特定健保組合
    if 31 <= h <= 34:
        return 3   # 共済
    if h == 39:
        return 6   # 後期高齢

    return None


def derive_clinic_code(toroku_kigo_no: str) -> str:
    """
    登録記号番号から施術機関コードを導出（暫定ルール）。
    例: "契2804440-0-0" → "2804440-0-0"

    ルール: 先頭の「協」または「契」の1文字のみ除去。ハイフンはそのまま保持。
    ★ 公式一次資料での完全確認未完了。現時点の暫定運用。
      （詳細: docs/JREC-01_申請書様式運用メモ.md §4 U2 参照）
    """
    s = (toroku_kigo_no or "").strip()
    if s and s[0] in ("協", "契"):
        s = s[1:]
    return s  # ハイフン保持


def calc_age_at_end_of_month(birthday, ym: str):
    """
    対象月末日時点の年齢を計算する。

    Args:
        birthday: 生年月日（date / "YYYY-MM-DD" 文字列 / スプレッドシートの日付値）
        ym: 対象月 "YYYY-MM"
    Returns:
        int（年齢） or None（計算不能）
    """
    if not birthday or not ym:
        return None
    bd = parse_date(birthday)
    if not bd:
        return None
    try:
        parts = ym.split("-")
        y, m_num = int(parts[0]), int(parts[1])
        # 対象月末日
        if m_num == 12:
            eom = date(y + 1, 1, 1) - timedelta(days=1)
        else:
            eom = date(y, m_num + 1, 1) - timedelta(days=1)
    except (ValueError, IndexError):
        return None
    age = eom.year - bd.year
    if (eom.month, eom.day) < (bd.month, bd.day):
        age -= 1
    return age if age >= 0 else None


def derive_honkeku_cell(row1: dict) -> str | None:
    """
    U5本家区分: 患者情報から書込セル番地を返す。

    判定ロジック（優先順位順）:
    1. 保険種別=6（後期高齢）→ DH8（8.高一）基本。7割給付（負担3割）のみ DH12（0.高7）
       ★制度確定（2026-03-20）: 本人/家族区分は使わない。給付割合は U6 側で表現する。
    2. 6歳未満（就学前）→ DB10（4.六歳）
    3. 70〜74歳 + 一部負担金割合=2 → DH8（8.高一）
    4. 70〜74歳 + 一部負担金割合=3 → DH12（0.高7）
    5. 70〜74歳 + 割合不明 → DH8（高一で安全側）
    6. 75歳以上 → DH8（8.高一）基本。7割給付（負担3割）のみ DH12（0.高7）
       ★保険種別が6以外でも75歳超は後期高齢者として高一/高7 で判定する
    7. 70歳未満 + 続柄=本人 → DB8（2.本人）
    8. 70歳未満 + 続柄その他 → DB12（6.家族）

    生年月日がない場合: 続柄のみで本人/家族を判定（年齢区分はスキップ）。
    """
    # 保険種別: 数値(6)も名称文字列("後期高齢")も数値に正規化
    # GASマスタは"協会けんぽ"等の文字列で保存されているため、文字列→数値マップが必要
    _INS_TYPE_NAME_MAP = {
        "協会けんぽ": 1, "組合": 2, "共済": 3, "国保": 4, "退職": 5, "後期高齢": 6,
    }
    ins_type_raw = row1.get("保険種別")
    ins_type = safe_int(ins_type_raw)
    if ins_type is None and ins_type_raw:
        ins_type = _INS_TYPE_NAME_MAP.get(str(ins_type_raw).strip())
    if ins_type is None:
        ins_type = detect_insurance_type(str(row1.get("保険者番号") or "")) or 0
    ins_type = ins_type or 0

    relation = str(row1.get("続柄") or "").strip()
    burden   = safe_int(row1.get("一部負担金割合")) or 0
    birthday = row1.get("患者生年月日")
    ym       = str(row1.get("対象月") or "")

    # 後期高齢者（保険種別=6）→ 高一 基本。7割給付（負担3割）のみ 高7
    if ins_type == 6:
        return "DH12" if burden == 3 else "DH8"

    age = calc_age_at_end_of_month(birthday, ym)

    # 6歳未満（就学前）
    if age is not None and age < 6:
        return "DB10"

    # 70〜74歳（前期高齢者）
    if age is not None and 70 <= age <= 74:
        if burden == 2:
            return "DH8"
        if burden == 3:
            return "DH12"
        return "DH8"  # 負担割合不明は安全側（高一=8割給付）

    # 75歳以上 → 後期高齢者扱い（保険種別=6と同じルール）
    if age is not None and age >= 75:
        return "DH12" if burden == 3 else "DH8"

    # 70歳未満（年齢不明含む）: 続柄で判定
    return "DB8" if relation == "本人" else "DB12"


def put_calendar_circles(ws, cell, visit_days):
    """
    M32の施術日カレンダーテキスト内で通院日の数字を丸数字に置換。

    元テキスト: "1   2   3   4   5   ...  31"
    visit_days: [3, 5, 10, 15, ...] のような通院日リスト

    ★置換は大きい数字(31)から順に行う（"1"が"10"の中の1を誤置換しないように）
    """
    text = ws[cell].value
    if not text:
        return

    # 大きい数字から順に置換
    days_sorted = sorted(set(visit_days), reverse=True)
    for day in days_sorted:
        if day < 1 or day > 31:
            continue
        circled = CIRCLED_NUMBERS.get(day, str(day))
        target = str(day)
        # テキスト内の該当数字を見つけて丸数字に置換
        # ただし "10" の中の "1" を置換しないよう、前後の文字を確認
        text = _replace_day_in_calendar(text, day, circled)

    ws[cell] = text


def _replace_day_in_calendar(text, day, replacement):
    """
    カレンダーテキスト内の日番号を丸数字に置換。
    前後がスペースまたは文字列の端であることを確認して誤置換を防ぐ。
    """
    target = str(day)
    tlen = len(target)
    result = []
    i = 0
    replaced = False
    while i < len(text):
        if not replaced and text[i:i+tlen] == target:
            # 前後チェック：前がスペースor先頭、後がスペースor末尾
            before_ok = (i == 0 or text[i-1] == ' ')
            after_ok = (i + tlen >= len(text) or text[i+tlen] == ' ')
            if before_ok and after_ok:
                result.append(replacement)
                i += tlen
                replaced = True
                continue
        result.append(text[i])
        i += 1
    return ''.join(result)


# ===== データ組み立て =====

def build_injury_rows(row1, row2):
    """負傷名行データ（最大4行）"""
    result = []
    # case1 部位1
    result.append({
        "name":       row1.get("負傷名1", ""),
        "injuryDate": row1.get("負傷年月日1", ""),
        "firstDate":  row1.get("初検年月日1", ""),
        "startDate":  row1.get("施術開始年月日1", ""),
        "endDate":    row1.get("施術終了年月日1", ""),
        "days":       row1.get("実日数1", ""),
        "tenki":      row1.get("転帰1", ""),
    })
    # case1 部位2
    result.append({
        "name":       row1.get("負傷名2", ""),
        "injuryDate": row1.get("負傷年月日2", ""),
        "firstDate":  row1.get("初検年月日2", ""),
        "startDate":  row1.get("施術開始年月日2", ""),
        "endDate":    row1.get("施術終了年月日2", ""),
        "days":       row1.get("実日数2", ""),
        "tenki":      row1.get("転帰2", ""),
    })
    if row2:
        # case2 部位1
        result.append({
            "name":       row2.get("負傷名1", ""),
            "injuryDate": row2.get("負傷年月日1", ""),
            "firstDate":  row2.get("初検年月日1", ""),
            "startDate":  row2.get("施術開始年月日1", ""),
            "endDate":    row2.get("施術終了年月日1", ""),
            "days":       row2.get("実日数1", ""),
            "tenki":      row2.get("転帰1", ""),
        })
        # case2 部位2
        result.append({
            "name":       row2.get("負傷名2", ""),
            "injuryDate": row2.get("負傷年月日2", ""),
            "firstDate":  row2.get("初検年月日2", ""),
            "startDate":  row2.get("施術開始年月日2", ""),
            "endDate":    row2.get("施術終了年月日2", ""),
            "days":       row2.get("実日数2", ""),
            "tenki":      row2.get("転帰2", ""),
        })
    return result


def build_shoryo_array(row1, row2):
    """施療料配列"""
    arr = [
        safe_num(row1.get("施療料1")) or 0,
        safe_num(row1.get("施療料2")) or 0,
    ]
    if row2:
        arr.append(safe_num(row2.get("施療料1")) or 0)
        arr.append(safe_num(row2.get("施療料2")) or 0)
    return arr


def build_part_detail_array(row1, row2):
    """部位別明細データ"""
    result = []

    def build_one(row, part_no):
        pfx = f"部位{part_no}_"
        koryo_amt = safe_num(row.get(f"{pfx}後療料_金額")) or 0
        subtotal = safe_num(row.get(f"{pfx}計")) or 0
        has_data = subtotal > 0 or koryo_amt > 0
        return {
            "has_data":   has_data,
            "tei_rate":   row.get(f"{pfx}逓減率"),
            "koryo_unit": safe_num(row.get(f"{pfx}後療料_単価")),
            "koryo_cnt":  safe_num(row.get(f"{pfx}後療料_回数")),
            "koryo_amt":  safe_num(row.get(f"{pfx}後療料_金額")),
            "cold_cnt":   safe_num(row.get(f"{pfx}冷罨法_回数")),
            "cold_amt":   safe_num(row.get(f"{pfx}冷罨法_金額")),
            "warm_cnt":   safe_num(row.get(f"{pfx}温罨法_回数")),
            "warm_amt":   safe_num(row.get(f"{pfx}温罨法_金額")),
            "elec_cnt":   safe_num(row.get(f"{pfx}電療_回数")),
            "elec_amt":   safe_num(row.get(f"{pfx}電療_金額")),
            "subtotal":   safe_num(row.get(f"{pfx}計")),
        }

    result.append(build_one(row1, 1))
    result.append(build_one(row1, 2))
    if row2:
        result.append(build_one(row2, 1))
        result.append(build_one(row2, 2))
    return result


# ===== バッチモード（NDJSON） =====

SCHEMA_VERSION = "3.0"

# case1に必須のキー
REQUIRED_CASE1_KEYS = ["患者ID", "対象月", "患者氏名", "当月合計", "窓口負担額", "請求金額"]

# Excel出力後の検証対象セル
VERIFY_CELLS = {
    "患者氏名": "E21",
    "合計桁1": "DP44",  # 合計の1の位（最低でもここに値があるべき）
}


def load_batch_ndjson(ndjson_path: str) -> dict:
    """
    NDJSONファイルを読み込む。
    1行目: メタデータ（_meta: true）
    2行目〜: 患者データ（patientId, case1, case2, visitDays）

    返り値: {"meta": {...}, "patients": [...]}
    """
    meta = None
    patients = []

    with open(ndjson_path, "r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as e:
                print(f"エラー: {line_no}行目のJSON解析に失敗: {e}")
                sys.exit(1)

            if obj.get("_meta"):
                meta = obj
            else:
                patients.append(obj)

    if meta is None:
        print("エラー: NDJSONにメタデータ行（_meta: true）がありません。")
        sys.exit(1)

    return {"meta": meta, "patients": patients}


def validate_batch(batch: dict):
    """
    バッチデータのバリデーション。問題があればsys.exit(1)で停止。

    チェック項目:
    1. schemaVersionが "3.0" であること
    2. patientCountとリスト件数の一致
    3. patientId重複チェック
    4. case1必須キー存在チェック
    5. case1["患者ID"]==patientId, case1["対象月"]==meta.month の整合性チェック
    """
    meta = batch["meta"]
    patients = batch["patients"]
    errors = []

    # 1. schemaVersion
    sv = str(meta.get("schemaVersion", ""))
    if sv != SCHEMA_VERSION:
        errors.append(f"schemaVersion不一致: 期待={SCHEMA_VERSION}, 実際={sv}")

    # 2. patientCount
    expected_count = meta.get("patientCount", 0)
    if expected_count != len(patients):
        errors.append(f"patientCount不一致: メタ={expected_count}, 実際={len(patients)}")

    # 3. patientId重複チェック
    seen_ids = {}
    for i, p in enumerate(patients):
        pid = p.get("patientId", "")
        if pid in seen_ids:
            errors.append(f"patientId重複: {pid}（行{seen_ids[pid]+2}と行{i+2}）")
        else:
            seen_ids[pid] = i

    month = meta.get("month", "")

    # 4 & 5. 各患者のcase1チェック
    for i, p in enumerate(patients):
        pid = p.get("patientId", f"[行{i+2}]")
        case1 = p.get("case1")
        if case1 is None:
            errors.append(f"{pid}: case1がnullです")
            continue

        # 4. 必須キー存在チェック
        for key in REQUIRED_CASE1_KEYS:
            val = case1.get(key)
            if val is None or val == "":
                errors.append(f"{pid}: case1に必須キー「{key}」がありません")

        # 5. 整合性チェック
        c1_pid = str(case1.get("患者ID", "")).strip()
        if c1_pid and c1_pid != pid:
            errors.append(f"{pid}: case1[\"患者ID\"]={c1_pid} がpatientIdと不一致")

        c1_month = str(case1.get("対象月", "")).strip()
        if c1_month and c1_month != month:
            errors.append(f"{pid}: case1[\"対象月\"]={c1_month} がmeta.month={month}と不一致")

    if errors:
        print("バリデーションエラー:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)

    print(f"バリデーション OK: {len(patients)}患者, schemaVersion={sv}, month={month}")


def verify_output(output_path: str, expected: dict) -> list:
    """
    出力Excelの必須セル検証。
    expected: {"患者氏名": "...", "当月合計": 1234, ...}

    返り値: 警告メッセージリスト（空なら問題なし）
    """
    warnings = []
    try:
        wb = load_workbook(output_path, read_only=True, data_only=True)
        ws = wb[TEMPLATE_SHEET]

        # 患者氏名
        name_val = ws["E21"].value
        if not name_val:
            warnings.append("E21（患者氏名）が空")
        elif expected.get("患者氏名") and str(name_val) != str(expected["患者氏名"]):
            warnings.append(f"E21（患者氏名）不一致: Excel={name_val}, JSON={expected['患者氏名']}")

        # 合計欄（DP44 = 1の位。合計>0なら少なくともここに値があるべき）
        total = safe_int(expected.get("当月合計"))
        if total and total > 0:
            dp44_val = ws["DP44"].value
            if dp44_val is None:
                warnings.append("DP44（合計1の位）が空（当月合計>0なのに）")

        # 保険者番号（CQ4 = 最上位桁）
        insurer = expected.get("保険者番号", "")
        if insurer:
            cq4_val = ws["CQ4"].value
            if cq4_val is None:
                warnings.append("CQ4（保険者番号1桁目）が空")

        wb.close()
    except Exception as e:
        warnings.append(f"検証中にエラー: {e}")

    return warnings


def batch_write(ndjson_path: str):
    """
    NDJSONファイルからバッチ転記を実行。
    """
    script_dir = Path(__file__).parent
    template = script_dir / TEMPLATE_FILE

    if not template.exists():
        print(f"テンプレートファイルが見つかりません: {template}")
        sys.exit(1)

    # 1. NDJSON読み込み
    print(f"NDJSON読み込み: {ndjson_path}")
    batch = load_batch_ndjson(ndjson_path)

    # 2. バリデーション
    validate_batch(batch)

    meta = batch["meta"]
    patients = batch["patients"]
    month = meta["month"]

    # 施術機関固定情報を meta から取得（NDJSON に含まれない場合は空でスキップ）
    clinic_info = {
        "prefectureNo": str(meta.get("prefectureNo") or ""),
        "torokuKigoNo": str(meta.get("torokuKigoNo") or ""),
    }

    # 3. 出力ディレクトリ作成
    out_dir = script_dir / "output" / month
    out_dir.mkdir(parents=True, exist_ok=True)

    # 4. 各患者の転記
    all_warnings = []
    success_count = 0

    for i, p in enumerate(patients):
        pid = p.get("patientId", "unknown")
        print(f"[{i+1}/{len(patients)}] {pid} ...", end=" ")

        json_data = {
            "case1": p.get("case1"),
            "case2": p.get("case2"),
            "visitDays": p.get("visitDays", []),
        }

        out_path = out_dir / f"申請書_{pid}_{month}.xlsx"

        try:
            write_application(str(template), json_data, str(out_path), clinic_info=clinic_info)
            success_count += 1

            # 機械検証
            expected = {
                "患者氏名": (p.get("case1") or {}).get("患者氏名", ""),
                "当月合計": (p.get("case1") or {}).get("当月合計", 0),
                "保険者番号": (p.get("case1") or {}).get("保険者番号", ""),
            }
            warns = verify_output(str(out_path), expected)
            if warns:
                all_warnings.append((pid, warns))
                print(f"警告あり ({len(warns)}件)")
            else:
                print("OK")
        except Exception as e:
            print(f"エラー: {e}")
            all_warnings.append((pid, [f"転記エラー: {e}"]))

    # 5. サマリー
    print(f"\n{'='*50}")
    print(f"一括転記完了: {success_count}/{len(patients)} 件")
    print(f"出力先: {out_dir}")

    if all_warnings:
        print(f"\n警告 ({len(all_warnings)} 患者):")
        for pid, warns in all_warnings:
            for w in warns:
                print(f"  [{pid}] {w}")
    else:
        print("検証: 全件パス")


# ===== ダウンロードフォルダ自動検出 =====

DEFAULT_DOWNLOAD_DIR = r"G:\マイドライブ\ダウンロード"


def find_latest_ndjson() -> tuple[str | None, str]:
    """
    NDJSONファイルを自動検出。
    探索先: 環境変数 V3_BATCH_DOWNLOAD_DIR > DEFAULT_DOWNLOAD_DIR
    返り値: (ファイルパス or None, 探索したフォルダパス)
    """
    search_dir = os.environ.get("V3_BATCH_DOWNLOAD_DIR", DEFAULT_DOWNLOAD_DIR)

    if not Path(search_dir).exists():
        return None, search_dir

    pattern = str(Path(search_dir) / "transfer_batch_*.ndjson")
    files = glob_mod.glob(pattern)
    if not files:
        return None, search_dir

    # 更新日時が最新のファイルを選ぶ
    latest = max(files, key=os.path.getmtime)
    return latest, search_dir


# ===== サーバー向けAPI（Cloud Run 用） =====

def write_application_to_bytes(template_path: str, json_data: dict, clinic_info: dict = None) -> bytes:
    """
    write_application() の in-memory 版。
    一時ファイルを経由して xlsx bytes を返す（保存なし）。
    clinic_info: 施術機関固定情報（batch_write_from_string から渡す）。None でもよい。
    """
    import io as _io
    import tempfile as _tmp
    tmp = _tmp.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()
    try:
        write_application(template_path, json_data, tmp_path, clinic_info=clinic_info)
        with open(tmp_path, "rb") as f:
            return f.read()
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def load_batch_ndjson_string(ndjson_str: str) -> dict:
    """
    NDJSON 文字列を解析して {"meta": {...}, "patients": [...]} を返す。
    エラー時は ValueError / JSONDecodeError を raise（sys.exit しない）。
    """
    meta = None
    patients = []
    for line_no, line in enumerate(ndjson_str.splitlines(), 1):
        line = line.strip()
        if not line:
            continue
        obj = json.loads(line)  # JSONDecodeError は呼び出し元に伝播
        if obj.get("_meta"):
            meta = obj
        else:
            patients.append(obj)
    if meta is None:
        raise ValueError("NDJSONにメタデータ行（_meta: true）がありません")
    return {"meta": meta, "patients": patients}


def validate_batch_safe(batch: dict) -> list:
    """
    バリデーション。エラーリストを返す（sys.exit しない）。
    空リストなら OK。
    """
    meta = batch["meta"]
    patients = batch["patients"]
    errors = []

    sv = str(meta.get("schemaVersion", ""))
    if sv != SCHEMA_VERSION:
        errors.append(f"schemaVersion不一致: 期待={SCHEMA_VERSION}, 実際={sv}")

    expected_count = meta.get("patientCount", 0)
    if expected_count != len(patients):
        errors.append(f"patientCount不一致: メタ={expected_count}, 実際={len(patients)}")

    seen_ids: dict = {}
    for i, p in enumerate(patients):
        pid = p.get("patientId", "")
        if pid in seen_ids:
            errors.append(f"patientId重複: {pid}")
        else:
            seen_ids[pid] = i

    month = meta.get("month", "")
    for i, p in enumerate(patients):
        pid = p.get("patientId", f"[行{i+2}]")
        case1 = p.get("case1")
        if case1 is None:
            errors.append(f"{pid}: case1がnullです")
            continue
        for key in REQUIRED_CASE1_KEYS:
            val = case1.get(key)
            if val is None or val == "":
                errors.append(f"{pid}: case1に必須キー「{key}」がありません")
        c1_month = str(case1.get("対象月", "")).strip()
        if c1_month and month and c1_month != month:
            errors.append(f"{pid}: case1[\"対象月\"]={c1_month} がmeta.month={month}と不一致")

    return errors


def verify_output_bytes(xlsx_bytes: bytes, expected: dict) -> list:
    """
    xlsx bytes から検証する（verify_output の in-memory 版）。
    """
    import io as _io
    warnings_list = []
    try:
        wb = load_workbook(_io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws = wb[TEMPLATE_SHEET]

        name_val = ws["E21"].value
        if not name_val:
            warnings_list.append("E21（患者氏名）が空")
        elif expected.get("患者氏名") and str(name_val) != str(expected["患者氏名"]):
            warnings_list.append(f"E21（患者氏名）不一致: Excel={name_val}, JSON={expected['患者氏名']}")

        total = safe_int(expected.get("当月合計"))
        if total and total > 0:
            if ws["DP44"].value is None:
                warnings_list.append("DP44（合計1の位）が空（当月合計>0なのに）")

        if expected.get("保険者番号") and ws["CQ4"].value is None:
            warnings_list.append("CQ4（保険者番号1桁目）が空")

        wb.close()
    except Exception as e:
        warnings_list.append(f"検証中にエラー: {e}")
    return warnings_list


def batch_write_from_string(ndjson_str: str, template_path: str = None) -> list:
    """
    NDJSON 文字列を受け取り、患者ごとの xlsx bytes を返す（Cloud Run 向け）。

    返り値: [
      {"patientId": str, "fileName": str, "content": bytes, "warnings": list},
      ...  # エラー患者は "error" キーを追加し content=None
    ]
    """
    if template_path is None:
        template_path = str(Path(__file__).parent / TEMPLATE_FILE)

    batch = load_batch_ndjson_string(ndjson_str)
    errors = validate_batch_safe(batch)
    if errors:
        raise ValueError("バリデーションエラー:\n" + "\n".join(f"  - {e}" for e in errors))

    meta = batch["meta"]
    month = meta["month"]

    # 施術機関固定情報を meta から取得
    clinic_info = {
        "prefectureNo": str(meta.get("prefectureNo") or ""),
        "torokuKigoNo": str(meta.get("torokuKigoNo") or ""),
    }

    results = []

    for p in batch["patients"]:
        pid = p.get("patientId", "unknown")
        json_data = {
            "case1": p.get("case1"),
            "case2": p.get("case2"),
            "visitDays": p.get("visitDays", []),
        }
        file_name = f"申請書_{pid}_{month}.xlsx"
        try:
            xlsx_bytes = write_application_to_bytes(template_path, json_data, clinic_info=clinic_info)
            expected = {
                "患者氏名": (p.get("case1") or {}).get("患者氏名", ""),
                "当月合計": (p.get("case1") or {}).get("当月合計", 0),
                "保険者番号": (p.get("case1") or {}).get("保険者番号", ""),
            }
            warns = verify_output_bytes(xlsx_bytes, expected)
            results.append({
                "patientId": pid,
                "fileName": file_name,
                "content": xlsx_bytes,
                "warnings": warns,
            })
        except Exception as e:
            results.append({
                "patientId": pid,
                "fileName": file_name,
                "content": None,
                "warnings": [f"生成エラー: {e}"],
                "error": str(e),
            })

    return results


# ===== メイン =====
def main():
    script_dir = Path(__file__).parent
    template = script_dir / TEMPLATE_FILE

    if not template.exists():
        print(f"テンプレートファイルが見つかりません: {template}")
        sys.exit(1)

    # バッチモード
    if len(sys.argv) >= 2 and sys.argv[1] == "--batch":
        if len(sys.argv) >= 3:
            # 明示的にファイル指定
            ndjson_path = sys.argv[2]
        else:
            # 自動検出
            ndjson_path, search_dir = find_latest_ndjson()
            if ndjson_path is None:
                print(f"エラー: transfer_batch_*.ndjson が見つかりません。")
                print(f"  探索先: {search_dir}")
                print(f"  環境変数 V3_BATCH_DOWNLOAD_DIR で探索先を変更できます。")
                sys.exit(1)
            print(f"自動検出: {ndjson_path}")

        if not Path(ndjson_path).exists():
            print(f"NDJSONファイルが見つかりません: {ndjson_path}")
            sys.exit(1)
        batch_write(ndjson_path)
        return

    # 単一患者モード（従来互換）
    if len(sys.argv) >= 2:
        json_path = Path(sys.argv[1])
    else:
        # 対話モード
        json_input = input("JSONファイルパスを入力（または直接JSONを貼り付け）:\n").strip()
        if json_input.startswith("{"):
            # 直接JSON
            data = json.loads(json_input)
            patient_id = (data.get("case1") or {}).get("患者ID", "unknown")
            ym = (data.get("case1") or {}).get("対象月", "unknown")
            output_name = f"申請書_{patient_id}_{ym}.xlsx"
            output_path = script_dir / output_name
            write_application(str(template), data, str(output_path))
            return
        json_path = Path(json_input)

    if not json_path.exists():
        print(f"JSONファイルが見つかりません: {json_path}")
        sys.exit(1)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    patient_id = (data.get("case1") or {}).get("患者ID", "unknown")
    ym = (data.get("case1") or {}).get("対象月", "unknown")
    output_name = f"申請書_{patient_id}_{ym}.xlsx"
    output_path = script_dir / output_name

    write_application(str(template), data, str(output_path))


if __name__ == "__main__":
    main()
