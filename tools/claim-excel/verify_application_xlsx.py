#!/usr/bin/env python3
"""
申請書 Excel 出力の自動確認スクリプト

生成された申請書 Excel ファイルを読み込み、以下を確認する:
1. 負傷名欄 (行26-30) の空行詰め状態
2. 施術日カレンダー (行32) の○印の有無
3. 金額欄の値
4. シート名

使い方:
  python verify_application_xlsx.py <xlsx_file>
  python verify_application_xlsx.py  (最新の xlsx を自動検出)
"""

import sys
import glob as glob_mod
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    sys.exit(1)

TEMPLATE_SHEET = "新　様式第5号"

# 負傷名欄のセル（write_application.py の INJURY_ROWS と同一）
INJURY_NAME_CELLS = ["E26", "E27", "E28", "E29", "E30"]

# 施術日カレンダー開始（行32 M32 から4列おき）
CALENDAR_START_COL = 13  # M列 (0-indexed: 12, 1-indexed: 13)
CALENDAR_ROW = 32
CALENDAR_STEP = 4

# 金額欄
AMOUNT_CELLS = {
    "合計": None,   # 複数セル結合のため実際の値は別途確認
    "一部負担金": None,
    "請求金額": None,
}


def find_latest_xlsx():
    """最新の申請書 xlsx を自動検出"""
    pattern = str(Path(__file__).parent / "申請書_*.xlsx")
    files = sorted(glob_mod.glob(pattern), key=os.path.getmtime, reverse=True)
    if not files:
        return None
    return files[0]


def get_cell_value(ws, cell_ref):
    """セルの値を取得（結合セルを考慮）"""
    cell = ws[cell_ref]
    return cell.value


def check_injury_rows(ws):
    """負傷名欄（行26-30）の状態を確認"""
    print("\n=== 負傷名欄 ===")
    rows = []
    for cell_ref in INJURY_NAME_CELLS:
        val = get_cell_value(ws, cell_ref)
        label = f"({len(rows) + 1})" if val else "   "
        print(f"  {cell_ref}: {label} {val!r}")
        rows.append(val)

    # 空行の位置チェック
    non_empty = [(i, v) for i, v in enumerate(rows) if v]
    empty = [(i, v) for i, v in enumerate(rows) if not v]

    print(f"\n  有効: {len(non_empty)} 行, 空: {len(empty)} 行")

    # 空行が詰まっているか確認
    if non_empty:
        last_non_empty_idx = non_empty[-1][0]
        gap = any(
            not rows[i]
            for i in range(last_non_empty_idx + 1)
            if i < len(rows) and i not in [x[0] for x in non_empty]
        )
        if gap:
            print("  ⚠️  WARNING: 有効行の間に空行があります（詰め不足）")
        else:
            print("  ✅ OK: 有効行が連続しています（空行詰め正常）")

    return rows


def check_calendar(ws):
    """施術日カレンダー（行32）の○印を確認"""
    print("\n=== 施術日カレンダー（行32） ===")
    circles = []
    for day in range(1, 32):
        col = CALENDAR_START_COL + (day - 1) * CALENDAR_STEP
        try:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col)
            cell_ref = f"{col_letter}{CALENDAR_ROW}"
            val = get_cell_value(ws, cell_ref)
            if val and "○" in str(val):
                circles.append(day)
        except Exception:
            pass

    if circles:
        print(f"  ✅ 施術日: {circles} （{len(circles)}日分の○）")
    else:
        print("  ⚠️  施術日カレンダーに○なし（または未実装）")

    return circles


def check_amounts(ws):
    """金額欄の確認"""
    print("\n=== 金額（参考）===")
    # 合計欄は複数セル結合で1桁ずつ入力されているため、
    # ここでは代表セルの値を出力するのみ
    check_cells = {
        "合計_右端": "DX44",    # 1円の位
        "一部負担_右端": "DX45",
        "請求金額_右端": "DX46",
    }
    for label, cell_ref in check_cells.items():
        try:
            val = get_cell_value(ws, cell_ref)
            print(f"  {cell_ref} ({label}): {val!r}")
        except Exception as e:
            print(f"  {cell_ref} ({label}): エラー - {e}")


def main():
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    else:
        xlsx_path = find_latest_xlsx()
        if not xlsx_path:
            print("申請書 xlsx ファイルが見つかりません。")
            print("引数でパスを指定してください: python verify_application_xlsx.py <file.xlsx>")
            sys.exit(1)
        print(f"最新ファイルを使用: {xlsx_path}")

    print(f"\n確認対象: {xlsx_path}")

    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"ファイルを開けません: {e}")
        sys.exit(1)

    print(f"シート一覧: {wb.sheetnames}")

    if TEMPLATE_SHEET not in wb.sheetnames:
        print(f"⚠️  シート「{TEMPLATE_SHEET}」が見つかりません")
        sys.exit(1)

    ws = wb[TEMPLATE_SHEET]
    print(f"シート「{TEMPLATE_SHEET}」を確認中...")

    check_injury_rows(ws)
    check_calendar(ws)
    check_amounts(ws)

    print("\n=== 確認完了 ===")
    print("目視確認が必要な項目:")
    print("  1. 負傷名欄の表示（(1)(2) に正しく入っているか）")
    print("  2. 施術日カレンダーの○（来院日に対応しているか）")
    print("  3. 転帰の丸囲み")
    print("  4. 金額が ¥3,053 相当か")
    print("  5. 印刷プレビューで1ページに収まるか")


if __name__ == "__main__":
    main()
